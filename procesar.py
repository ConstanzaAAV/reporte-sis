# -*- coding: utf-8 -*-
"""
Motor de procesamiento del 'REPORTE SIS' (Plan de Intervención / Mejor Niñez).
Recibe el PDF en memoria y devuelve el Excel (con datos + gráficos) también en memoria.

Función principal:  procesar_pdf(pdf_bytes) -> (xlsx_bytes, stats)

Reglas de negocio:
- Cada EVENTO se ancla en una fecha (dd-mm-aaaa) de la columna 'Fecha Evento'.
- Las descripciones largas CONTINÚAN en páginas siguientes (se unen solas).
- Se quitan las reimpresiones EXACTAS (duplicados de arrastre del reporte).
- VISITA REAL = combinación única (niño + fecha + técnico).
"""
import re
import io
import gc
import calendar
from collections import defaultdict, Counter

import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as col_l

TOL = 2.0
BOUNDS = [
    ("tipo_intervencion",   0,   134),
    ("nivel_intervencion",  134, 210),
    ("fecha_evento",        210, 272),
    ("tipo_evento",         272, 383),
    ("descripcion_evento",  383, 595),
    ("tecnico",             595, 705),
    ("f_actualizacion",     705, 100000),
]
COLS = [b[0] for b in BOUNDS]
DATE = re.compile(r"^\d{2}-\d{2}-\d{4}$")
MESES = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio",
         "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]


# ----------------------------- EXTRACCIÓN -----------------------------
def _col_of(x0):
    for name, lo, hi in BOUNDS:
        if lo <= x0 < hi:
            return name
    return None


def _reconstruct(ws):
    ws = sorted(ws, key=lambda w: (w["page"], w["top"]))
    lines, last = [], None
    for w in ws:
        if last is not None and w["page"] == last[0] and abs(w["top"] - last[1]) <= 4:
            lines[-1][1].append(w)
        else:
            lines.append([(w["page"], w["top"]), [w]])
        last = (w["page"], w["top"])
    out = []
    for _, lw in lines:
        lw = sorted(lw, key=lambda w: w["x0"])
        out.append(" ".join(w["text"] for w in lw))
    return " ".join(out).strip()


def _extract_name(words):
    jv = [w for w in words if w["text"] == "Joven"]
    if not jv:
        return ""
    top = jv[0]["top"]
    line = sorted([w for w in words if abs(w["top"] - top) <= 3.5], key=lambda w: w["x0"])
    colon = None
    for i, w in enumerate(line):
        if w["text"] == ":":
            colon = i
    if colon is None:
        return ""
    toks, skip = [], True
    for w in line[colon + 1:]:
        if skip and re.match(r"^[\d.,]+$", w["text"]):
            continue
        skip = False
        toks.append(w["text"])
    return " ".join(toks).strip()


def _new_event(name, page):
    ev = {c: [] for c in COLS}
    ev["nombre"] = name
    ev["pagina"] = page
    return ev


def _is_dup(a, b):
    if a["nombre"] != b["nombre"] or a["fecha_evento"] != b["fecha_evento"]:
        return False
    if a["tipo_evento"] != b["tipo_evento"] or a["tecnico"] != b["tecnico"]:
        return False
    da, db = a["descripcion_evento"], b["descripcion_evento"]
    s, l = (da, db) if len(da) <= len(db) else (db, da)
    return len(s) > 0 and l.startswith(s)


def extraer_eventos(pdf_file):
    records, current = [], None

    def finalize():
        nonlocal current
        if current is None:
            return
        rec = {c: _reconstruct(current[c]) for c in COLS}
        rec["nombre"] = current["nombre"]
        rec["pagina"] = current["pagina"]
        records.append(rec)

    with pdfplumber.open(pdf_file) as pdf:
        for pno, page in enumerate(pdf.pages, 1):
            words = page.extract_words()
            page.flush_cache()          # liberar memoria (clave para PDFs grandes)
            for w in words:
                w["page"] = pno
            if pno % 50 == 0:
                gc.collect()
            name = _extract_name(words)
            hdr = [w for w in words if w["text"] in ("Descripción", "Técnico") and w["top"] > 250]
            if not hdr:
                continue
            ht = min(w["top"] for w in hdr)
            data = [w for w in words if w["top"] > ht + 5]
            anchors = sorted(w["top"] for w in data
                             if _col_of(w["x0"]) == "fecha_evento" and DATE.match(w["text"]))
            if not anchors:
                if current is not None:
                    for w in data:
                        c = _col_of(w["x0"])
                        if c:
                            current[c].append(w)
                continue
            first = anchors[0]
            if current is not None:
                for w in data:
                    if w["top"] < first - TOL:
                        c = _col_of(w["x0"])
                        if c:
                            current[c].append(w)
            for k, a in enumerate(anchors):
                lo = a - TOL
                hi = (anchors[k + 1] - TOL) if k + 1 < len(anchors) else float("inf")
                finalize()
                current = _new_event(name, pno)
                for w in data:
                    if lo <= w["top"] < hi:
                        c = _col_of(w["x0"])
                        if c:
                            current[c].append(w)
        finalize()

    # des-duplicar reimpresiones exactas
    clean = []
    for r in records:
        if clean and _is_dup(clean[-1], r):
            if len(r["descripcion_evento"]) > len(clean[-1]["descripcion_evento"]):
                clean[-1] = r
        else:
            clean.append(r)
    # marcar 'revisar' (mismo niño/fecha/técnico/desc pero distinto tipo de evento)
    for r in clean:
        r["revisar"] = ""
    for i in range(1, len(clean)):
        a, b = clean[i - 1], clean[i]
        if (a["nombre"] == b["nombre"] and a["fecha_evento"] == b["fecha_evento"]
                and a["tecnico"] == b["tecnico"]
                and a["descripcion_evento"] == b["descripcion_evento"]
                and a["tipo_evento"] != b["tipo_evento"]):
            a["revisar"] = b["revisar"] = "REVISAR (posible repetición)"
    return clean


# ----------------------------- AUXILIARES -----------------------------
def _mes_dias(recs):
    mc = Counter()
    for r in recs:
        m = re.match(r"\d{2}-(\d{2})-(\d{4})", r["fecha_evento"])
        if m:
            mc[(int(m.group(2)), int(m.group(1)))] += 1   # (año, mes)
    if not mc:
        return ["%02d-01-2026" % d for d in range(1, 32)], ""
    (yr, mo), _ = mc.most_common(1)[0]
    nd = calendar.monthrange(yr, mo)[1]
    dias = ["%02d-%02d-%04d" % (d, mo, yr) for d in range(1, nd + 1)]
    return dias, "%s %d" % (MESES[mo], yr)


def _disp(t):
    t = t or ""
    if "PSICOSOCIAL" in t: return "Psicosocial"
    if "REDES" in t: return "Redes"
    if "JURÍDICO" in t: return "Jurídico"
    if "PARENTALES" in t: return "Comp. parentales"
    if "SALUD" in t: return "Salud/educación"
    if "CRISIS" in t: return "Crisis"
    if "REUNIFIC" in t: return "Reunificación"
    return "Sin tipo"


# ----------------------------- EXCEL + GRÁFICOS -----------------------------
def construir_excel(recs):
    dias, etiqueta = _mes_dias(recs)

    ti_counter = Counter(r["tipo_intervencion"] for r in recs)
    te_counter = Counter(r["tipo_evento"] for r in recs)
    tipos_int = [t for t, _ in ti_counter.most_common() if t != ""]
    has_blank_int = "" in ti_counter
    tipos_evt = [t for t, _ in te_counter.most_common() if t != ""]

    by_child = defaultdict(list)
    for r in recs:
        by_child[r["nombre"]].append(r)
    resumen = []
    for n, rs in by_child.items():
        filas = len(rs)
        vr = len(set((x["fecha_evento"], x["tecnico"]) for x in rs))
        resumen.append({
            "nombre": n, "filas": filas, "vr": vr,
            "ratio": round(filas / vr, 2) if vr else 0,
            "int": {t: sum(1 for x in rs if x["tipo_intervencion"] == t) for t in tipos_int},
            "sin": sum(1 for x in rs if x["tipo_intervencion"] == ""),
        })
    resumen.sort(key=lambda d: (-d["vr"], -d["filas"]))
    n_ninos = len(resumen)
    vr_global = len(set((r["nombre"], r["fecha_evento"], r["tecnico"]) for r in recs))

    # datos del gráfico circular (tipo de intervención resumido)
    _intc = [(t, ti_counter[t]) for t in tipos_int] + ([("", ti_counter[""])] if has_blank_int else [])
    _intc.sort(key=lambda x: -x[1])
    _otros = sum(c for _, c in _intc[4:])
    pie_data = [(_disp(t), c) for t, c in _intc[:4]] + ([("Otros", _otros)] if _otros > 0 else [])

    # textos del recuadro de preguntas clave
    _avg = vr_global / n_ninos if n_ninos else 0
    _factor = len(recs) / vr_global if vr_global else 0
    _t3 = resumen[:3]
    _intens = sum(1 for d in resumen if d["vr"] >= 10)
    _pocos = sum(1 for d in resumen if d["vr"] <= 2)
    _tevt = te_counter.most_common(1)[0] if te_counter else ("", 0)
    _sin = ti_counter.get("", 0)
    _p3 = [(_disp(t), 100 * ti_counter[t] / len(recs)) for t in tipos_int[:3]]
    while len(_p3) < 3:
        _p3.append(("—", 0))

    def _nm(s):
        return (s[:22] + "…") if len(s) > 22 else s

    insights = [
        ("¿Cuántos niños y atenciones?",
         "%d niños · %d visitas reales (de %d registros)" % (n_ninos, vr_global, len(recs))),
        ("¿Visitas reales promedio por niño?",
         "%.1f visitas por niño en el período" % _avg),
        ("¿El registro infla las cifras?",
         "Sí: %d registros = solo %d visitas reales (factor %.2f)" % (len(recs), vr_global, _factor)),
        ("¿Qué tipo de trabajo predomina?",
         "%s (%.0f%%); le siguen %s (%.0f%%) y %s (%.0f%%)" %
         (_p3[0][0], _p3[0][1], _p3[1][0], _p3[1][1], _p3[2][0], _p3[2][1])),
        ("¿A quién se atendió más?",
         "%s (%d), %s (%d) y %s (%d) visitas" %
         (_nm(_t3[0]["nombre"]), _t3[0]["vr"], _nm(_t3[1]["nombre"]), _t3[1]["vr"],
          _nm(_t3[2]["nombre"]), _t3[2]["vr"]) if len(_t3) >= 3 else "—"),
        ("¿Atención concentrada o repartida?",
         "%d niños con 10+ visitas (intensivos); %d con ≤2 (poca actividad)" % (_intens, _pocos)),
        ("¿Gestión más frecuente?",
         "%s (%d veces)" % (_tevt[0].capitalize(), _tevt[1])),
        ("¿Calidad del registro?",
         "%d registros (%.0f%%) sin clasificar (sin tipo)" % (_sin, 100 * _sin / len(recs) if recs else 0)),
    ]

    ORANGE = PatternFill("solid", fgColor="C55A11")
    RED = PatternFill("solid", fgColor="C00000")
    YEL = PatternFill("solid", fgColor="FFF2A8")
    BOLDW = Font(bold=True, color="FFFFFF")

    wb = openpyxl.Workbook()

    # ---- EVENTOS ----
    ev = wb.active
    ev.title = "Eventos"
    heads = ["Página", "Nombre", "Tipo Intervención", "Fecha Evento", "Tipo Evento",
             "Descripción Evento", "Técnico", "Revisar"]
    keys = ["pagina", "nombre", "tipo_intervencion", "fecha_evento", "tipo_evento",
            "descripcion_evento", "tecnico", "revisar"]
    ev.append(heads)
    for c in ev[1]:
        c.font = BOLDW; c.fill = RED
    for r in recs:
        ev.append([r[k] for k in keys])
        if r["revisar"]:
            for c in ev[ev.max_row]:
                c.fill = YEL
    for col, w in zip("ABCDEFGH", [8, 32, 25, 12, 28, 80, 25, 22]):
        ev.column_dimensions[col].width = w
    for row in ev.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=(c.column_letter in ("B", "C", "E", "F")))
    ev.freeze_panes = "A2"
    ev.auto_filter.ref = "A1:H%d" % ev.max_row

    # ---- RESUMEN POR NIÑO ----
    rs = wb.create_sheet("Resumen por niño")
    head2 = (["Niño", "Intervenciones (filas)", "VISITAS REALES", "Interv. por visita"]
             + tipos_int + (["(sin tipo)"] if has_blank_int else []))
    rs.append(head2)
    for c in rs[1]:
        c.font = BOLDW; c.fill = ORANGE
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for d in resumen:
        row = [d["nombre"], d["filas"], d["vr"], d["ratio"]] + [d["int"][t] for t in tipos_int]
        if has_blank_int:
            row.append(d["sin"])
        rs.append(row)
    rs.column_dimensions["A"].width = 34
    for i in range(2, len(head2) + 1):
        rs.column_dimensions[col_l(i)].width = 13
    rs.row_dimensions[1].height = 60
    rs.freeze_panes = "B2"
    rs.auto_filter.ref = "A1:%s%d" % (col_l(len(head2)), rs.max_row)
    NR = rs.max_row

    # ---- AUX (oculta) ----
    ax = wb.create_sheet("Aux")

    def write_block(col, header, items):
        ax.cell(row=1, column=col, value=header[0])
        ax.cell(row=1, column=col + 1, value=header[1])
        r = 2
        for label, val in items:
            ax.cell(row=r, column=col, value=label)
            ax.cell(row=r, column=col + 1, value=val)
            r += 1
        return col, col + 1, 2, r - 1

    SEL = "Panel!$B$3"
    items_int = [(t, "=COUNTIFS(Eventos!$B:$B,%s,Eventos!$C:$C,$A%d)" % (SEL, 2 + i))
                 for i, t in enumerate(tipos_int)]
    ci_l, ci_v, ci0, _ = write_block(1, ("Tipo Intervención", "Cuenta"), items_int)
    ci1 = 1 + len(tipos_int)
    if has_blank_int:
        ci1 += 1
        ax.cell(row=ci1, column=1, value="(sin tipo)")
        ax.cell(row=ci1, column=2, value="=COUNTIF(Eventos!$B:$B,%s)-SUM(B2:B%d)" % (SEL, 1 + len(tipos_int)))

    tipos_evt_panel = [t for t, _ in te_counter.most_common(25) if t != ""][:20]
    items_evt = [(t, "=COUNTIFS(Eventos!$B:$B,%s,Eventos!$E:$E,$D%d)" % (SEL, 2 + i))
                 for i, t in enumerate(tipos_evt_panel)]
    ce_l, ce_v, ce0, ce1 = write_block(4, ("Tipo Evento", "Cuenta"), items_evt)

    items_dia = [(d, "=COUNTIFS(Eventos!$B:$B,%s,Eventos!$D:$D,$G%d)" % (SEL, 2 + i))
                 for i, d in enumerate(dias)]
    cd_l, cd_v, cd0, cd1 = write_block(7, ("Día", "Intervenciones"), items_dia)

    top = resumen[:25]
    ct_l, ct_v, ct0, ct1 = write_block(10, ("Niño", "Visitas reales"),
                                       [(d["nombre"], d["vr"]) for d in top])

    eg = te_counter.most_common(20)
    eg = [("(sin tipo)" if k == "" else k, v) for k, v in eg]
    eg_l, eg_v, eg0, eg1 = write_block(16, ("Tipo Evento", "Total"), eg)

    cc_l, cc_v, cc0, cc1 = write_block(19, ("Métrica", "Cantidad"),
                                       [("Intervenciones (filas)", len(recs)), ("Visitas reales", vr_global)])
    cp_l, cp_v, cp0, cp1 = write_block(22, ("Tipo (resumido)", "Total"), pie_data)
    ax.sheet_state = "hidden"

    # ---- PANEL (interactivo) ----
    pn = wb.create_sheet("Panel")
    pn["A1"] = "PANEL POR NIÑO"
    pn["A1"].font = Font(bold=True, size=16, color="C55A11")
    pn["A3"] = "Elige un niño:"
    pn["A3"].font = Font(bold=True)
    pn["B3"] = resumen[0]["nombre"] if resumen else ""
    pn["B3"].fill = YEL; pn["B3"].font = Font(bold=True)
    dv = DataValidation(type="list", formula1="='Resumen por niño'!$A$2:$A$%d" % NR, allow_blank=False)
    pn.add_data_validation(dv)
    dv.add(pn["B3"])

    VL = "VLOOKUP($B$3,'Resumen por niño'!$A$2:$D$%d" % NR

    def kpi(cl, cv, label, formula):
        pn[cl] = label; pn[cl].font = Font(bold=True)
        pn[cv] = formula; pn[cv].font = Font(bold=True, size=20, color="C00000")

    kpi("A5", "A6", "VISITAS REALES (contactos)", "=%s,3,FALSE)" % VL)
    kpi("C5", "C6", "Intervenciones registradas", "=%s,2,FALSE)" % VL)
    kpi("E5", "E6", "Intervenciones por visita", "=%s,4,FALSE)" % VL)
    for cc in ("A", "C", "E"):
        pn.column_dimensions[cc].width = 26

    def bar(ws, anchor, title, lc, vc, r0, r1, kind="col", h=8, w=16, datalabels=False):
        ch = BarChart(); ch.type = kind; ch.title = title; ch.legend = None
        ch.height = h; ch.width = w; ch.varyColors = False
        ch.add_data(Reference(ax, min_col=vc, min_row=1, max_row=r1), titles_from_data=True)
        ch.set_categories(Reference(ax, min_col=lc, min_row=r0, max_row=r1))
        ch.x_axis.delete = False; ch.y_axis.delete = False
        ch.x_axis.tickLblPos = "low"
        if datalabels:
            d = DataLabelList(); d.showVal = True
            d.showSerName = False; d.showCatName = False
            d.showLegendKey = False; d.showPercent = False; d.showBubbleSize = False
            ch.dataLabels = d
        ws.add_chart(ch, anchor)

    bar(pn, "A9", "Actividad por día (nº intervenciones) — %s" % etiqueta, cd_l, cd_v, cd0, cd1, "col", 8, 24)
    bar(pn, "A28", "Tipos de intervención del niño", ci_l, ci_v, ci0, ci1, "col", 9, 15)
    bar(pn, "K28", "Tipos de evento del niño (top 20 más frecuentes)", ce_l, ce_v, ce0, ce1, "bar", 12, 18)

    # ---- GENERAL ----
    gn = wb.create_sheet("General")
    gn["A1"] = "VISTA GENERAL — TODOS LOS NIÑOS" + ((" · " + etiqueta) if etiqueta else "")
    gn["A1"].font = Font(bold=True, size=16, color="C55A11")
    gn["A3"] = "Intervenciones registradas: %d   |   Visitas reales: %d   |   Niños: %d" % (len(recs), vr_global, n_ninos)
    gn["A3"].font = Font(bold=True)

    thin = Side(style="thin", color="C55A11")
    boxb = Border(left=thin, right=thin, top=thin, bottom=thin)
    gn.merge_cells("A5:J5")
    hc = gn["A5"]
    hc.value = "📌 PREGUNTAS CLAVE QUE RESPONDE ESTA VISTA"
    hc.font = Font(bold=True, color="FFFFFF"); hc.fill = ORANGE
    hc.alignment = Alignment(horizontal="left", vertical="center")
    rr = 6
    for q, a in insights:
        gn.merge_cells("A%d:C%d" % (rr, rr))
        gn.merge_cells("D%d:J%d" % (rr, rr))
        qc = gn.cell(row=rr, column=1, value=q)
        qc.font = Font(bold=True, size=10); qc.alignment = Alignment(wrap_text=True, vertical="center")
        ac = gn.cell(row=rr, column=4, value=a)
        ac.font = Font(size=10); ac.alignment = Alignment(wrap_text=True, vertical="center")
        gn.row_dimensions[rr].height = 30
        rr += 1
    for row in gn.iter_rows(min_row=5, max_row=rr - 1, min_col=1, max_col=10):
        for cell in row:
            cell.border = boxb

    pie = PieChart()
    pie.title = "¿Qué tipo de trabajo predomina?"
    pie.add_data(Reference(ax, min_col=cp_v, min_row=1, max_row=cp1), titles_from_data=True)
    pie.set_categories(Reference(ax, min_col=cp_l, min_row=cp0, max_row=cp1))
    pdl = DataLabelList(); pdl.showPercent = True; pdl.showVal = False; pdl.showCatName = False
    pie.dataLabels = pdl
    pie.height = 10.5; pie.width = 12.5
    gn.add_chart(pie, "L5")

    bar(gn, "A26", "Top 25 niños por VISITAS REALES", ct_l, ct_v, ct0, ct1, "bar", 15, 18)
    bar(gn, "M26", "Intervenciones (filas) vs Visitas reales", cc_l, cc_v, cc0, cc1, "col", 8, 10, datalabels=True)
    bar(gn, "A58", "Top 20 tipos de evento", eg_l, eg_v, eg0, eg1, "bar", 12, 18)

    # ---- POR TÉCNICO (carga de trabajo) ----
    by_tec = defaultdict(list)
    for r in recs:
        by_tec[r["tecnico"]].append(r)
    tec_rows = []
    for t, rr_ in by_tec.items():
        inter = len(rr_)
        vis = len(set((x["nombre"], x["fecha_evento"]) for x in rr_))
        na = len(set(x["nombre"] for x in rr_))
        tec_rows.append((t, inter, vis, na, round(inter / vis, 2) if vis else 0))
    tec_rows.sort(key=lambda x: -x[2])   # por visitas reales, desc

    tw = wb.create_sheet("Por técnico")
    tw["A1"] = "POR TÉCNICO — carga de trabajo" + ((" · " + etiqueta) if etiqueta else "")
    tw["A1"].font = Font(bold=True, size=15, color="C55A11")
    headT = ["Técnico", "Intervenciones (filas)", "Visitas reales", "Niños atendidos", "Interv. por visita"]
    for j, h in enumerate(headT, start=1):
        c = tw.cell(row=3, column=j, value=h)
        c.font = BOLDW; c.fill = ORANGE
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    tw.row_dimensions[3].height = 32
    for i, (t, inter, vis, na, ratio) in enumerate(tec_rows):
        rw = 4 + i
        tw.cell(row=rw, column=1, value=t)
        tw.cell(row=rw, column=2, value=inter)
        tw.cell(row=rw, column=3, value=vis)
        tw.cell(row=rw, column=4, value=na)
        tw.cell(row=rw, column=5, value=ratio)
    last_t = 3 + len(tec_rows)
    tw.column_dimensions["A"].width = 34
    for col in "BCDE":
        tw.column_dimensions[col].width = 18
    tw.freeze_panes = "A4"
    tw.auto_filter.ref = "A3:E%d" % last_t

    cht = BarChart(); cht.type = "bar"; cht.grouping = "clustered"; cht.varyColors = False
    cht.title = "Intervenciones vs visitas reales por técnico"
    cht.height = 0.9 * len(tec_rows) + 4; cht.width = 20
    cht.add_data(Reference(tw, min_col=2, max_col=3, min_row=3, max_row=last_t), titles_from_data=True)
    cht.set_categories(Reference(tw, min_col=1, min_row=4, max_row=last_t))
    cht.x_axis.delete = False; cht.y_axis.delete = False
    tw.add_chart(cht, "G3")

    # ---- GLOSARIO (primera hoja, "léeme") ----
    glos = [
        ("H", "QUÉ ES ESTE ARCHIVO", ""),
        ("T", "Resumen del mes", "Convierte el Reporte SIS (Plan de Intervención · Mejor Niñez) de un mes en datos ordenados y gráficos. Sirve para ver cuánto y cómo se atendió a cada niño, niña o adolescente."),
        ("H", "CONCEPTOS CLAVE", ""),
        ("T", "Registro (o intervención)", "Cada fila de la hoja 'Eventos'. Es una acción anotada en el sistema (una coordinación, una visita, una sesión, un informe, etc.). Una misma atención puede generar VARIOS registros."),
        ("T", "Visita REAL", "Un contacto/atención efectiva. Se define como la combinación única NIÑO + FECHA + TÉCNICO. Si un mismo profesional anota 5 cosas el mismo día para el mismo niño = 1 visita real. Si dos profesionales distintos lo atienden el mismo día = 2 visitas reales."),
        ("T", "¿Por qué dos números (registros y visitas)?", "El sistema permite anotar una sola visita en varias intervenciones. Contar registros INFLA la cantidad real de atenciones; las 'visitas reales' muestran cuántas veces se atendió de verdad."),
        ("T", "Interv. por visita", "Promedio de registros por visita real (Intervenciones ÷ Visitas reales). Cerca de 1 = poco subdividido; mientras más alto, más se anotó cada visita en muchos registros."),
        ("H", "LAS HOJAS DE ESTE EXCEL", ""),
        ("T", "Glosario", "Esta hoja: explica qué significa cada cosa."),
        ("T", "Eventos", "Todos los registros del mes, uno por fila (datos limpios y ordenados, con filtros activados arriba)."),
        ("T", "Resumen por niño", "Una fila por niño/a con sus totales: intervenciones, visitas reales, promedio, y cuántas de cada tipo de intervención."),
        ("T", "Panel", "Tablero interactivo: elige un niño en la lista desplegable (celda amarilla) y se actualizan sus gráficos (actividad por día, tipos de intervención y de evento)."),
        ("T", "General", "Vista de todo el programa: recuadro con preguntas clave, gráfico de tipos de trabajo (torta), ranking de niños por visitas reales y tipos de evento más frecuentes."),
        ("T", "Por técnico", "Carga de trabajo de cada profesional: cuántas intervenciones (registros) y cuántas visitas reales hizo, cuántos niños atendió y su promedio."),
        ("H", "COLUMNAS DE LA HOJA 'EVENTOS'", ""),
        ("T", "Página", "Página del PDF original donde empieza ese registro (para poder ubicarlo)."),
        ("T", "Nombre", "Niño, niña o adolescente atendido (apellidos + nombres)."),
        ("T", "Tipo Intervención", "Categoría general del trabajo: Psicosocial, Complementaria de Redes, Apoyo Jurídico, Competencias Parentales, etc."),
        ("T", "Fecha Evento", "Fecha en que ocurrió la acción (día-mes-año)."),
        ("T", "Tipo Evento", "Detalle específico de la acción: visita domiciliaria, sesión individual, coordinación con redes, elaboración de informe, etc."),
        ("T", "Descripción Evento", "Texto completo de lo que se hizo / se registró."),
        ("T", "Técnico", "Profesional que realizó y registró la acción."),
        ("T", "Revisar", "Marca 'REVISAR' (fila amarilla) en registros que parecen repetidos: misma descripción y técnico pero distinto Tipo Evento. Conviene revisarlos a ojo."),
        ("H", "CÓMO SE CALCULAN LOS NÚMEROS", ""),
        ("T", "Conteo por tipo (ej. Psicosocial = 9)", "Cuenta cuántas FILAS de ese niño tienen ese Tipo Intervención. Es por REGISTRO, no por visita real: 9 registros pueden corresponder a menos visitas (si varios fueron el mismo día y técnico)."),
        ("T", "Visitas reales", "Cantidad de combinaciones únicas NIÑO + FECHA + TÉCNICO."),
        ("T", "(sin tipo)", "Registros que en el PDF venían sin Tipo Intervención (el campo estaba vacío)."),
        ("T", "Duplicados", "El reporte a veces reimprime un registro al cambiar de página; esas copias exactas se eliminan automáticamente."),
        ("H", "NOTAS IMPORTANTES", ""),
        ("T", "Datos sensibles", "Contiene información de menores. Guárdalo en un lugar seguro y no lo compartas por canales abiertos."),
        ("T", "Ábrelo en Microsoft Excel", "Para ver bien los gráficos y el panel interactivo, ábrelo en Excel (en otros visores pueden verse distintos)."),
        ("T", "Período", "Este archivo corresponde a un mes. Para otro mes, genera uno nuevo con la herramienta."),
    ]
    gl = wb.create_sheet("Glosario")
    gl.sheet_view.showGridLines = False
    gl.column_dimensions["A"].width = 32
    gl.column_dimensions["B"].width = 98
    gl.merge_cells("A1:B1")
    gl["A1"] = "📘 GLOSARIO — Cómo leer este archivo"
    gl["A1"].font = Font(bold=True, size=16, color="C55A11")
    grow = 3
    for kind, term, dfn in glos:
        if kind == "H":
            gl.merge_cells("A%d:B%d" % (grow, grow))
            c = gl.cell(row=grow, column=1, value=term)
            c.font = Font(bold=True, color="FFFFFF"); c.fill = ORANGE
            c.alignment = Alignment(vertical="center")
            gl.row_dimensions[grow].height = 20
        else:
            t = gl.cell(row=grow, column=1, value=term)
            t.font = Font(bold=True, size=10); t.alignment = Alignment(wrap_text=True, vertical="top")
            d = gl.cell(row=grow, column=2, value=dfn)
            d.font = Font(size=10); d.alignment = Alignment(wrap_text=True, vertical="top")
            lines = (len(dfn) + 95) // 96
            gl.row_dimensions[grow].height = 15 * max(1, lines) + 3
        grow += 1
    # mover Glosario al frente y dejarlo como hoja activa
    wb._sheets.remove(gl)
    wb._sheets.insert(0, gl)
    wb.active = 0

    wb.calculation.fullCalcOnLoad = True
    buf = io.BytesIO()
    wb.save(buf)
    stats = {"ninos": n_ninos, "filas": len(recs), "visitas": vr_global, "mes": etiqueta}
    return buf.getvalue(), stats


def procesar_pdf(pdf_bytes):
    """PDF (bytes) -> (xlsx_bytes, stats dict)."""
    recs = extraer_eventos(io.BytesIO(pdf_bytes))
    return construir_excel(recs)
